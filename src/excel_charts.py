"""
Excel Chart Generator for Construction Dashboard
Creates additional Excel files with charts and pivot tables
"""

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import PatternFill, Font, Alignment
import os

def create_excel_with_charts():
    """Create Excel file with embedded charts and pivot analysis"""
    
    try:
        # Load data
        schedule_df = pd.read_csv('data/schedule_data.csv')
        cost_df = pd.read_csv('data/cost_data.csv')
        safety_df = pd.read_csv('data/safety_data.csv')
        quality_df = pd.read_csv('data/quality_data.csv')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets with charts
        create_schedule_charts_sheet(wb, schedule_df)
        create_cost_charts_sheet(wb, cost_df)
        create_safety_charts_sheet(wb, safety_df)
        create_executive_summary_sheet(wb, schedule_df, cost_df, safety_df, quality_df)
        
        # Save workbook
        chart_file = "Construction_Dashboard_Charts.xlsx"
        wb.save(chart_file)
        
        print(f"✅ Excel charts file created: {chart_file}")
        return chart_file
        
    except FileNotFoundError:
        print("❌ CSV data files not found. Run data_generator.py first.")
        return None
    except Exception as e:
        print(f"❌ Error creating charts: {str(e)}")
        return None

def create_schedule_charts_sheet(wb, schedule_df):
    """Create schedule performance charts"""
    
    ws = wb.create_sheet("Schedule Analysis")
    
    # Add data to worksheet
    ws['A1'] = 'Week'
    ws['B1'] = 'Planned Progress %'
    ws['C1'] = 'Actual Progress %'
    ws['D1'] = 'SPI'
    
    for i, row in schedule_df.iterrows():
        ws[f'A{i+2}'] = row['Week']
        ws[f'B{i+2}'] = row['Planned_Progress_Pct']
        ws[f'C{i+2}'] = row['Actual_Progress_Pct']
        ws[f'D{i+2}'] = row['SPI']
    
    # Create progress chart
    chart1 = LineChart()
    chart1.title = "Schedule Performance - Progress Tracking"
    chart1.style = 13
    chart1.y_axis.title = 'Progress (%)'
    chart1.x_axis.title = 'Project Week'
    
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(schedule_df)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(schedule_df)+1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    ws.add_chart(chart1, "F2")
    
    # Create SPI chart
    chart2 = LineChart()
    chart2.title = "Schedule Performance Index (SPI)"
    chart2.style = 12
    chart2.y_axis.title = 'SPI'
    chart2.x_axis.title = 'Project Week'
    
    spi_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=len(schedule_df)+1)
    chart2.add_data(spi_data, titles_from_data=True)
    chart2.set_categories(cats)
    
    ws.add_chart(chart2, "F18")

def create_cost_charts_sheet(wb, cost_df):
    """Create cost performance charts"""
    
    ws = wb.create_sheet("Cost Analysis")
    
    # Add data
    ws['A1'] = 'Week'
    ws['B1'] = 'Cumulative Budget'
    ws['C1'] = 'Cumulative Spent'
    ws['D1'] = 'CPI'
    ws['E1'] = 'Cost Variance'
    
    for i, row in cost_df.iterrows():
        ws[f'A{i+2}'] = row['Week']
        ws[f'B{i+2}'] = row['Cumulative_Budget']
        ws[f'C{i+2}'] = row['Cumulative_Spent']
        ws[f'D{i+2}'] = row['CPI']
        ws[f'E{i+2}'] = row['Cost_Variance']
    
    # Budget vs Actual chart
    chart1 = LineChart()
    chart1.title = "Budget vs Actual Spending"
    chart1.style = 13
    chart1.y_axis.title = 'Cost ($)'
    chart1.x_axis.title = 'Project Week'
    
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(cost_df)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(cost_df)+1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    ws.add_chart(chart1, "G2")
    
    # CPI chart
    chart2 = LineChart()
    chart2.title = "Cost Performance Index (CPI)"
    chart2.style = 12
    chart2.y_axis.title = 'CPI'
    chart2.x_axis.title = 'Project Week'
    
    cpi_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=len(cost_df)+1)
    chart2.add_data(cpi_data, titles_from_data=True)
    chart2.set_categories(cats)
    
    ws.add_chart(chart2, "G18")

def create_safety_charts_sheet(wb, safety_df):
    """Create safety performance charts"""
    
    ws = wb.create_sheet("Safety Analysis")
    
    # Add data
    ws['A1'] = 'Week'
    ws['B1'] = 'Days Since Incident'
    ws['C1'] = 'TRIR'
    ws['D1'] = 'Near Miss Count'
    
    for i, row in safety_df.iterrows():
        ws[f'A{i+2}'] = row['Week']
        ws[f'B{i+2}'] = row['Days_Since_Last_Incident']
        ws[f'C{i+2}'] = row['TRIR']
        ws[f'D{i+2}'] = row['Near_Miss_Count']
    
    # Safety trend chart
    chart1 = LineChart()
    chart1.title = "Safety Performance Trends"
    chart1.style = 13
    chart1.y_axis.title = 'Days / Rate'
    chart1.x_axis.title = 'Project Week'
    
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(safety_df)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(safety_df)+1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    ws.add_chart(chart1, "F2")

def create_executive_summary_sheet(wb, schedule_df, cost_df, safety_df, quality_df):
    """Create executive summary with key metrics"""
    
    ws = wb.create_sheet("Executive Summary", 0)  # Insert as first sheet
    
    # Title and formatting
    ws['A1'] = 'CONSTRUCTION PROJECT DASHBOARD'
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    
    # Project status
    ws['A3'] = 'PROJECT STATUS OVERVIEW'
    ws['A3'].font = Font(size=14, bold=True)
    
    # Latest metrics
    latest_schedule = schedule_df.iloc[-1]
    latest_cost = cost_df.iloc[-1]
    latest_safety = safety_df.iloc[-1]
    latest_quality = quality_df.iloc[-1]
    
    # KPI Summary
    metrics = [
        ('Current Progress', f"{latest_schedule['Actual_Progress_Pct']:.1f}%"),
        ('Schedule Performance Index (SPI)', f"{latest_schedule['SPI']:.3f}"),
        ('Cost Performance Index (CPI)', f"{latest_cost['CPI']:.3f}"),
        ('Days Ahead/Behind', f"{latest_schedule['Days_Variance']:+.1f}"),
        ('Budget Spent', f"${latest_cost['Cumulative_Spent']:,.0f}"),
        ('Budget Remaining', f"${5000000 - latest_cost['Cumulative_Spent']:,.0f}"),
        ('Days Since Incident', f"{latest_safety['Days_Since_Last_Incident']}"),
        ('TRIR', f"{latest_safety['TRIR']:.2f}"),
        ('Inspection Pass Rate', f"{latest_quality['Inspection_Pass_Rate_Pct']:.1f}%"),
        ('Open Punch Items', f"{latest_quality['Punch_List_Items']}")
    ]
    
    # Add metrics to sheet
    ws['A5'] = 'Key Performance Indicator'
    ws['B5'] = 'Current Value'
    ws['C5'] = 'Status'
    
    for i, (metric, value) in enumerate(metrics):
        row = i + 6
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        
        # Status indicators
        if 'SPI' in metric or 'CPI' in metric:
            float_val = float(value)
            ws[f'C{row}'] = '🟢 Good' if float_val >= 1.0 else '🔴 Poor'
        elif 'Days Since' in metric:
            days = int(value)
            ws[f'C{row}'] = '🟢 Good' if days > 30 else '🟡 Warning'
        else:
            ws[f'C{row}'] = '📊 Monitor'
    
    # Project timeline
    ws['A17'] = 'PROJECT TIMELINE'
    ws['A17'].font = Font(size=14, bold=True)
    
    timeline_data = [
        ('Project Start', '2024-01-15'),
        ('Current Date', '2024-10-20'),
        ('Planned Completion', '2024-12-20'),
        ('Days Elapsed', f"{(pd.to_datetime('2024-10-20') - pd.to_datetime('2024-01-15')).days}"),
        ('Days Remaining', f"{(pd.to_datetime('2024-12-20') - pd.to_datetime('2024-10-20')).days}"),
        ('Percent Time Elapsed', f"{((pd.to_datetime('2024-10-20') - pd.to_datetime('2024-01-15')).days / (pd.to_datetime('2024-12-20') - pd.to_datetime('2024-01-15')).days * 100):.1f}%")
    ]
    
    for i, (item, value) in enumerate(timeline_data):
        row = i + 18
        ws[f'A{row}'] = item
        ws[f'B{row}'] = value

def create_pivot_analysis():
    """Create Excel file with pivot table analysis"""
    
    try:
        # This would require xlwings or similar for full pivot functionality
        # For now, create a summary analysis file
        
        cost_breakdown = pd.read_csv('data/cost_breakdown.csv')
        
        with pd.ExcelWriter('Construction_Pivot_Analysis.xlsx', engine='openpyxl') as writer:
            cost_breakdown.to_excel(writer, sheet_name='Cost Analysis', index=False)
            
            # Add summary calculations
            summary_data = {
                'Category': ['Total Budget', 'Total Actual', 'Total Variance', 'Percent Over Budget'],
                'Amount': [
                    cost_breakdown['Budget'].sum(),
                    cost_breakdown['Actual'].sum(),
                    cost_breakdown['Variance'].sum(),
                    f"{(cost_breakdown['Actual'].sum() - cost_breakdown['Budget'].sum()) / cost_breakdown['Budget'].sum() * 100:.1f}%"
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        print("✅ Pivot analysis file created: Construction_Pivot_Analysis.xlsx")
        
    except Exception as e:
        print(f"❌ Error creating pivot analysis: {str(e)}")

def main():
    """Generate Excel charts and analysis files"""
    
    print("📊 Creating Excel Charts and Analysis Files")
    print("=" * 45)
    
    # Create charts file
    chart_file = create_excel_with_charts()
    
    # Create pivot analysis
    create_pivot_analysis()
    
    if chart_file:
        print(f"\n✅ Excel files ready:")
        print(f"📈 Charts & Analysis: {chart_file}")
        print(f"📊 Pivot Analysis: Construction_Pivot_Analysis.xlsx")
        
        print(f"\n🎯 Chart Features:")
        print("• Schedule performance line charts")
        print("• Cost tracking with variance analysis")  
        print("• Safety trend monitoring")
        print("• Executive summary dashboard")
        print("• Formatted KPI indicators")

if __name__ == "__main__":
    main()